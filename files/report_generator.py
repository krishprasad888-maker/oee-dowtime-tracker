"""
report_generator.py
-------------------
Exports a formatted Excel report with:
  - Sheet 1: Executive summary (OEE by line)
  - Sheet 2: Full shift-level detail
  - Sheet 3: Downtime Pareto
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = "outputs"

# Colour palette (hex without #)
GREEN  = "1D9E75"
AMBER  = "BA7517"
RED    = "E24B4A"
HEADER = "2C2C2A"
WHITE  = "FFFFFF"
LGREY  = "F1EFE8"


def _oee_fill(value: float) -> PatternFill:
    """Return a green/amber/red fill depending on OEE value."""
    if value >= 85:
        color = "C0DD97"   # green-light
    elif value >= 70:
        color = "FAC775"   # amber-light
    else:
        color = "F09595"   # red-light
    return PatternFill("solid", fgColor=color)


def _header_style():
    return Font(bold=True, color=WHITE, size=11)


def _thin_border():
    side = Side(style="thin", color="D3D1C7")
    return Border(left=side, right=side, top=side, bottom=side)


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)


def generate_report(df_detail: pd.DataFrame,
                    df_by_line: pd.DataFrame,
                    pareto: pd.Series,
                    filename: str = "OEE_Report.xlsx"):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

        # ── Sheet 1: Line summary ──────────────────────────────────────────
        df_by_line.to_excel(writer, sheet_name="Summary by Line", index=False)

        # ── Sheet 2: Full detail ───────────────────────────────────────────
        cols_out = [
            "date", "shift", "line", "planned_production_time_min",
            "downtime_min", "downtime_reason",
            "availability", "performance", "quality", "oee"
        ]
        df_detail[cols_out].round(2).to_excel(
            writer, sheet_name="Shift Detail", index=False
        )

        # ── Sheet 3: Pareto ────────────────────────────────────────────────
        pareto_df = pareto.reset_index()
        pareto_df.columns = ["Downtime Reason", "Total Minutes"]
        pareto_df["Cumulative %"] = (
            pareto_df["Total Minutes"].cumsum() / pareto_df["Total Minutes"].sum() * 100
        ).round(1)
        pareto_df.to_excel(writer, sheet_name="Downtime Pareto", index=False)

    # ── Post-process: apply styles via openpyxl ───────────────────────────
    wb = load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Style header row
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=HEADER)
            cell.font = _header_style()
            cell.alignment = Alignment(horizontal="center")

        # Alternate row shading + OEE colour coding
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            bg = LGREY if row_idx % 2 == 0 else WHITE
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = _thin_border()
                cell.alignment = Alignment(horizontal="center")

        # OEE column colour coding on Summary and Detail sheets
        if sheet_name in ("Summary by Line", "Shift Detail"):
            header_vals = [c.value for c in ws[1]]
            if "avg_oee" in header_vals:
                oee_col = header_vals.index("avg_oee") + 1
            elif "oee" in header_vals:
                oee_col = header_vals.index("oee") + 1
            else:
                oee_col = None

            if oee_col:
                for row in ws.iter_rows(min_row=2, min_col=oee_col, max_col=oee_col):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.fill = _oee_fill(cell.value)
                            cell.font = Font(bold=True, size=11)

        _autofit(ws)

    wb.save(filepath)
    print(f"  Saved: {filepath}")
    return filepath
